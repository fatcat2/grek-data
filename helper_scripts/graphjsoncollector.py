from openpyxl import load_workbook
import json
wb = load_workbook(filename = '2019GreekDataPurdue.xlsx')
sheet_ranges = wb['Overall Rankings']

list = []

for row in sheet_ranges.iter_rows(min_row=2, min_col=1, max_row=66, max_col=13):
	a = []
	for cell in row:
		a.append(cell.value)
	list.append({"x": a[9], "y": a[1], "name": a[0]})
text_file = open("new.txt", "w")
text_file.write(json.dumps(list))
text_file.close()