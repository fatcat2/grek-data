from openpyxl import load_workbook
import json
wb = load_workbook(filename = '2018sFSCL.xlsx')
sheet_ranges = wb['Overall Rankings']

list = []

for row in sheet_ranges.iter_rows(min_row=2, min_col=1, max_row=90, max_col=9):
	a = []
	for cell in row:
		a.append(cell.value)
	list.append({"name": a[0],"semester_gpa": a[1],"cum_gpa": a[2],"rank": a[3],"nm_gpa": a[4],"new_mem": a[5],"total_mem": a[6],"service_hrs": a[7],"phil": a[8]})
text_file = open("Output.txt", "w")
text_file.write(json.dumps(list))
text_file.close()